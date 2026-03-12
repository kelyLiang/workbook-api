from __future__ import annotations

import io
import math
import re
import uuid
from copy import deepcopy
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

app = FastAPI(title="Workbook Parse API", version="1.2.2")

# =========================
# 内存存储（单机版）
# =========================
FILE_STORE: Dict[str, bytes] = {}
PARSED_WORKBOOK_STORE: Dict[str, Dict[str, Any]] = {}
MODEL_STORE: Dict[str, Dict[str, Any]] = {}
EXECUTION_STORE: Dict[str, Dict[str, Any]] = {}
DATASET_STORE: Dict[str, Dict[str, Any]] = {}


# =========================
# 基础工具函数
# =========================
def gen_request_id(prefix: str = "req_parse") -> str:
    return f"{prefix}_{uuid.uuid4().hex[:8]}"


def gen_model_id(prefix: str = "model") -> str:
    return f"{prefix}_{uuid.uuid4().hex[:8]}"


def gen_plan_id(prefix: str = "plan") -> str:
    return f"{prefix}_{uuid.uuid4().hex[:8]}"


def gen_job_id(prefix: str = "job") -> str:
    return f"{prefix}_{uuid.uuid4().hex[:8]}"


def gen_dataset_id(prefix: str = "ds") -> str:
    now = datetime.now().strftime("%Y%m%d")
    return f"{prefix}_{now}_{uuid.uuid4().hex[:6]}"


def gen_file_id(prefix: str = "file") -> str:
    return f"{prefix}_{uuid.uuid4().hex[:6]}"


def safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def normalize_cell_value(v: Any) -> Any:
    if v is None:
        return None

    if isinstance(v, (pd.Timestamp, datetime)):
        if pd.isna(v):
            return None
        return v.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")

    if isinstance(v, (np.integer,)):
        return int(v)

    if isinstance(v, (np.floating, float)):
        try:
            if pd.isna(v) or math.isinf(float(v)):
                return None
        except Exception:
            return None
        return float(v)

    if isinstance(v, (np.bool_,)):
        return bool(v)

    try:
        if pd.isna(v):
            return None
    except Exception:
        pass

    return v


def to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {k: to_jsonable(v) for k, v in value.items()}

    if isinstance(value, list):
        return [to_jsonable(v) for v in value]

    if isinstance(value, tuple):
        return [to_jsonable(v) for v in value]

    return normalize_cell_value(value)


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", "", safe_str(name)).lower()


def is_empty_row(row: List[Any]) -> bool:
    return all(cell is None or safe_str(cell) == "" for cell in row)


def is_blank_name(name: Any) -> bool:
    return safe_str(name) == ""


def truthy_flag(v: Any) -> bool:
    s = safe_str(v).lower()
    return s in {"1", "y", "yes", "true", "是", "阶梯", "阶梯提点"}


def ensure_object_column(df: pd.DataFrame, col: str) -> None:
    if col not in df.columns:
        df[col] = pd.Series([None] * len(df), index=df.index, dtype="object")
    else:
        if df[col].dtype != "object":
            df[col] = df[col].astype("object")


# =========================
# 字段角色识别工具
# =========================
def is_effective_start_field(name: str) -> bool:
    n = normalize_name(name)
    return any(k in n for k in ["生效", "开始", "起始", "start", "from"])


def is_effective_end_field(name: str) -> bool:
    n = normalize_name(name)
    return any(k in n for k in ["失效", "结束", "截止", "end", "to"])


def is_threshold_min_field(name: str) -> bool:
    n = normalize_name(name)
    return any(k in n for k in ["下限", "最小", "min", "lower"])


def is_threshold_max_field(name: str) -> bool:
    n = normalize_name(name)
    return any(k in n for k in ["上限", "最大", "max", "upper"])


def is_time_like_field(name: str) -> bool:
    n = normalize_name(name)
    return any(k in n for k in ["年月", "日期", "时间", "month", "date", "期间", "统计月"])


def is_ladder_flag_field(name: str) -> bool:
    n = normalize_name(name)
    return any(k in n for k in ["是否阶梯", "阶梯提点", "是否阶梯提点", "ladder", "isladder"])


# =========================
# Excel 解析工具
# =========================
def detect_header_and_rows(ws: Worksheet) -> Tuple[List[str], List[List[Any]]]:
    """
    简单假设：首个非空行是表头，其后为数据区
    """
    all_rows = list(ws.iter_rows(values_only=True))
    header: List[str] = []
    data_rows: List[List[Any]] = []

    for idx, row in enumerate(all_rows):
        row_list = list(row)
        if not is_empty_row(row_list):
            header = [safe_str(c) for c in row_list]
            data_rows = [list(r) for r in all_rows[idx + 1:] if not is_empty_row(list(r))]
            break

    return header, data_rows


def preview_values(data_rows: List[List[Any]], col_idx: int, n: int = 2) -> List[Any]:
    vals = []
    for row in data_rows:
        if col_idx < len(row):
            v = row[col_idx]
            if v is not None and safe_str(v) != "":
                vals.append(normalize_cell_value(v))
        if len(vals) >= n:
            break
    return vals


def infer_data_type(values: List[Any]) -> str:
    if not values:
        return "string"

    non_empty = [v for v in values if v is not None and safe_str(v) != ""]
    if not non_empty:
        return "string"

    date_like_count = 0
    num_count = 0

    for v in non_empty:
        if isinstance(v, datetime):
            date_like_count += 1
            continue

        s = safe_str(v)
        if re.match(r"^\d{4}[-/]\d{1,2}$", s) or re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$", s):
            date_like_count += 1
            continue

        try:
            float(v)
            num_count += 1
        except Exception:
            pass

    total = len(non_empty)
    if date_like_count / total >= 0.6:
        return "date"
    if num_count / total >= 0.6:
        return "number"
    return "string"


def infer_semantic_type(col_name: str, data_type: str) -> str:
    name = normalize_name(col_name)

    if any(k in name for k in ["年月", "日期", "时间", "month", "date", "期间", "统计月"]):
        return "time"
    if any(k in name for k in ["金额", "收入", "奖金", "提点率", "提点系数", "率", "下限", "上限", "base", "metric"]):
        return "metric"
    if any(k in name for k in ["工号", "uid", "id", "编码", "编号", "key"]):
        return "identifier"
    return "dimension"


def infer_sheet_role(sheet_name: str, header: List[str], data_rows: List[List[Any]]) -> Tuple[str, float]:
    name = normalize_name(sheet_name)
    clean_header = [h for h in header if safe_str(h)]

    if "逻辑" in name or "说明" in name or "分析" in name or "instruction" in name:
        return "instruction", 0.97

    if "mapping" in name or "关系" in name or "lookup" in name or "rule" in name:
        if any(is_threshold_min_field(h) or is_threshold_max_field(h) for h in clean_header):
            return "rule", 0.90
        if any(is_effective_start_field(h) or is_effective_end_field(h) for h in clean_header):
            return "rule", 0.88
        return "dimension", 0.92

    large_rows = len(data_rows) >= 50
    has_time = any(is_time_like_field(h) for h in clean_header)
    has_metric = any(("金额" in h or "收入" in h or "奖金" in h or "基数" in h) for h in clean_header)

    if large_rows and has_time and has_metric:
        return "fact", 0.96

    return "unknown", 0.60


def infer_grain_candidates(header: List[str]) -> List[List[str]]:
    clean_header = [h for h in header if safe_str(h)]
    header_norm_map = {normalize_name(h): h for h in clean_header}
    candidates = []

    preferred = []
    for key in ["统计年月", "年月", "工号", "uid", "收入类型", "产品线名称", "产品类型名称", "提点产品线名称"]:
        nk = normalize_name(key)
        if nk in header_norm_map:
            preferred.append(header_norm_map[nk])

    if preferred:
        candidates.append(preferred)

    if not candidates:
        dims = []
        for h in clean_header:
            nh = normalize_name(h)
            if any(k in nh for k in ["工号", "uid", "id", "类型", "名称", "年月", "日期"]):
                dims.append(h)
        if dims:
            candidates.append(dims[:4])

    return candidates


def summarize_instruction_sheet(data_rows: List[List[Any]], max_chars: int = 300) -> str:
    texts = []
    for row in data_rows[:100]:
        row_text = " ".join([safe_str(x) for x in row if safe_str(x)])
        if row_text:
            texts.append(row_text)

    merged = "；".join(texts)
    if not merged:
        return ""

    if len(merged) > max_chars:
        return merged[:max_chars] + "..."
    return merged


def build_sheet_profile(ws: Worksheet, preview_rows: int) -> Dict[str, Any]:
    header, data_rows = detect_header_and_rows(ws)

    row_count = len(data_rows)
    column_count = len(header)
    role, confidence = infer_sheet_role(ws.title, header, data_rows)
    grain_candidates = infer_grain_candidates(header)

    columns = []
    for idx, col_name in enumerate(header):
        if not safe_str(col_name):
            continue

        sample_vals = preview_values(data_rows[:preview_rows], idx, 2)
        data_type = infer_data_type(sample_vals if sample_vals else [])
        semantic_type = infer_semantic_type(col_name, data_type)

        columns.append({
            "name": col_name,
            "data_type": data_type,
            "semantic_type": semantic_type,
            "sample_values": sample_vals
        })

    return {
        "sheet_name": ws.title,
        "row_count": row_count,
        "column_count": column_count,
        "candidate_role": role,
        "role_confidence": round(confidence, 2),
        "grain_candidates": grain_candidates,
        "columns": columns,
        "_header": header,
        "_data_rows": data_rows
    }


def find_same_named_columns(from_header: List[str], to_header: List[str]) -> List[Tuple[str, str]]:
    to_norm_map = {normalize_name(c): c for c in to_header if safe_str(c)}
    matches = []
    for col in from_header:
        if not safe_str(col):
            continue
        n = normalize_name(col)
        if n in to_norm_map:
            matches.append((col, to_norm_map[n]))
    return matches


def infer_relationships(sheet_profiles: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    relationships = []
    rel_idx = 1

    facts = [s for s in sheet_profiles if s["candidate_role"] == "fact"]
    dims = [s for s in sheet_profiles if s["candidate_role"] in ("dimension", "rule")]

    for fact in facts:
        fact_header = [h for h in fact["_header"] if safe_str(h)]

        for dim in dims:
            dim_header = [h for h in dim["_header"] if safe_str(h)]
            matches = find_same_named_columns(fact_header, dim_header)

            if not matches:
                continue

            keys = [{"from_field": a, "to_field": b} for a, b in matches[:3]]
            matched_target_fields = {m[1] for m in matches}

            output_fields = [
                                c for c in dim_header
                                if safe_str(c) and c not in matched_target_fields
                            ][:3]

            confidence = min(0.75 + 0.08 * len(matches), 0.97)

            relationships.append({
                "relationship_id": f"rel_{rel_idx:03d}",
                "from_sheet": fact["sheet_name"],
                "to_sheet": dim["sheet_name"],
                "relationship_type": "lookup",
                "join_type": "left_join",
                "keys": keys,
                "output_fields": output_fields,
                "confidence": round(confidence, 2),
            })
            rel_idx += 1

    return relationships


def infer_candidate_rules(sheet_profiles: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rules = []
    rule_idx = 1

    def pick(header: List[str], *cands: str) -> Optional[str]:
        norm_map = {normalize_name(h): h for h in header if safe_str(h)}
        for c in cands:
            nc = normalize_name(c)
            if nc in norm_map:
                return norm_map[nc]
        return None

    for s in sheet_profiles:
        sheet_name = s["sheet_name"]
        header = [h for h in s["_header"] if safe_str(h)]
        if not header:
            continue

        start_field = None
        end_field = None
        min_field = None
        max_field = None
        output_field = None
        ladder_flag_field = None

        for h in header:
            hn = normalize_name(h)

            if start_field is None and any(k in hn for k in ["生效", "开始", "起始", "start", "from"]):
                start_field = h

            if end_field is None and any(k in hn for k in ["失效", "结束", "截止", "end", "to"]):
                end_field = h

            if min_field is None and any(k in hn for k in ["下限", "最小", "min", "lower"]):
                min_field = h

            if max_field is None and any(k in hn for k in ["上限", "最大", "max", "upper"]):
                max_field = h

            if output_field is None and any(k in hn for k in [
                "提点率", "提点系数", "系数", "费率", "比例", "rate", "factor", "coef"
            ]):
                output_field = h

            if ladder_flag_field is None and any(k in hn for k in [
                "是否阶梯", "阶梯提点", "是否阶梯提点", "ladder", "isladder"
            ]):
                ladder_flag_field = h

        match_keys = []
        excluded = {
            start_field, end_field, min_field, max_field, output_field,
            ladder_flag_field, None, ""
        }
        for h in header:
            if h in excluded:
                continue
            match_keys.append(h)

        has_effective = start_field is not None and end_field is not None
        has_threshold = min_field is not None or max_field is not None
        has_output = output_field is not None

        if has_output and (has_threshold or ladder_flag_field is not None):
            threshold_base_field = None
            for c in ["奖金计算基数", "业绩", "金额", "收入", "base"]:
                picked = pick(header, c)
                if picked:
                    threshold_base_field = picked
                    break

            rules.append({
                "rule_id": f"rule_{rule_idx:03d}",
                "rule_name": f"{normalize_name(sheet_name)}_productline_rate",
                "source_sheet": sheet_name,
                "rule_type": "productline_rate_lookup",
                "match_keys": match_keys[:3],
                "date_field": "统计年月",
                "effective_start_field": start_field,
                "effective_end_field": end_field,
                "ladder_flag_field": ladder_flag_field,
                "threshold_min_field": min_field,
                "threshold_max_field": max_field,
                "threshold_base_field": threshold_base_field,
                "output_field": output_field,
                "extra_output_fields": [x for x in [ladder_flag_field, min_field, max_field] if x],
                "confidence": 0.93 if ladder_flag_field else 0.88
            })
            rule_idx += 1
            continue

        if has_effective and has_output:
            rules.append({
                "rule_id": f"rule_{rule_idx:03d}",
                "rule_name": f"{normalize_name(sheet_name)}_lookup",
                "source_sheet": sheet_name,
                "rule_type": "effective_lookup",
                "match_keys": match_keys[:3],
                "date_field": "统计年月",
                "effective_start_field": start_field,
                "effective_end_field": end_field,
                "output_field": output_field,
                "confidence": 0.95
            })
            rule_idx += 1

    return rules


def cleanup_internal_fields(sheet_profiles: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    cleaned = []
    for s in sheet_profiles:
        copied = dict(s)
        copied.pop("_header", None)
        copied.pop("_data_rows", None)
        cleaned.append(copied)
    return cleaned


# =========================
# 模型构建工具
# =========================
def build_sheet_field_index(parsed_workbook: Dict[str, Any]) -> Dict[str, Set[str]]:
    index: Dict[str, Set[str]] = {}
    for sheet in parsed_workbook.get("sheets", []):
        sheet_name = sheet.get("sheet_name")
        fields = set()
        for col in sheet.get("columns", []):
            name = col.get("name")
            if name:
                fields.add(name)
        if sheet_name:
            index[sheet_name] = fields
    return index


def validate_sheet_exists(sheet_name: str, known_sheets: Set[str], errors: List[str], label: str) -> None:
    if sheet_name not in known_sheets:
        errors.append(f"{label} 不存在: {sheet_name}")


def validate_fields_exist(
        sheet_name: str,
        fields: List[str],
        field_index: Dict[str, Set[str]],
        errors: List[str],
        label: str,
) -> None:
    known_fields = field_index.get(sheet_name, set())
    for field in fields:
        if not field:
            continue
        if field not in known_fields:
            errors.append(f"{label} 字段不存在: sheet={sheet_name}, field={field}")


def infer_rule_action(rule_type: str) -> str:
    if rule_type in {"effective_lookup", "ladder_lookup", "lookup", "productline_rate_lookup"}:
        return rule_type
    return "rule"


def build_plan_steps(
        relationships: List[Dict[str, Any]],
        rules: List[Dict[str, Any]],
        formulas: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    steps: List[Dict[str, Any]] = []
    step_no = 1

    for rel in relationships:
        output_fields = [x for x in rel.get("output_fields", []) if safe_str(x)]
        target_name = "、".join(output_fields) if output_fields else rel.get("to_sheet", "关系映射")
        steps.append({
            "step_no": step_no,
            "action": "lookup",
            "name": f"映射{target_name}",
            "config": {
                "from_sheet": rel.get("from_sheet"),
                "to_sheet": rel.get("to_sheet"),
                "join_type": rel.get("join_type", "left_join"),
                "keys": rel.get("keys", []),
                "output_fields": output_fields,
            }
        })
        step_no += 1

    for rule in rules:
        rule_type = rule.get("rule_type", "rule")
        output_field = rule.get("output_field", "")
        display_name = output_field or rule.get("rule_name", "规则匹配")
        steps.append({
            "step_no": step_no,
            "action": infer_rule_action(rule_type),
            "name": f"匹配{display_name}",
            "config": rule
        })
        step_no += 1

    for formula in formulas:
        output_field = formula.get("output_field", "公式结果")
        steps.append({
            "step_no": step_no,
            "action": "formula",
            "name": f"计算{output_field}",
            "config": formula
        })
        step_no += 1

    return steps


def validate_model_request(
        payload: Dict[str, Any],
        parsed_workbook: Dict[str, Any]
) -> Dict[str, Any]:
    errors: List[str] = []
    warnings: List[str] = []

    workbook = parsed_workbook.get("workbook", {})
    known_sheets = set(workbook.get("sheet_names", []))
    field_index = build_sheet_field_index(parsed_workbook)

    main_fact_sheet = payload.get("main_fact_sheet")
    sheet_roles = payload.get("sheet_roles", [])
    relationships = payload.get("relationships", [])
    rules = payload.get("rules", [])
    formulas = payload.get("formulas", [])

    if not main_fact_sheet:
        errors.append("main_fact_sheet 不能为空")
    else:
        validate_sheet_exists(main_fact_sheet, known_sheets, errors, "main_fact_sheet")

    for item in sheet_roles:
        sheet_name = item.get("sheet_name")
        role = item.get("role")
        if not sheet_name:
            errors.append("sheet_roles 中存在缺少 sheet_name 的项")
            continue
        validate_sheet_exists(sheet_name, known_sheets, errors, "sheet_roles.sheet_name")
        if not role:
            errors.append(f"sheet_roles 缺少 role: {sheet_name}")

    if main_fact_sheet and sheet_roles:
        fact_role_ok = any(
            item.get("sheet_name") == main_fact_sheet and item.get("role") == "fact"
            for item in sheet_roles
        )
        if not fact_role_ok:
            warnings.append(f"main_fact_sheet={main_fact_sheet} 未在 sheet_roles 中标记为 fact")

    for i, rel in enumerate(relationships, start=1):
        from_sheet = rel.get("from_sheet")
        to_sheet = rel.get("to_sheet")
        keys = rel.get("keys", [])
        output_fields = [x for x in rel.get("output_fields", []) if safe_str(x)]

        if not from_sheet:
            errors.append(f"relationships[{i}] 缺少 from_sheet")
        else:
            validate_sheet_exists(from_sheet, known_sheets, errors, f"relationships[{i}].from_sheet")

        if not to_sheet:
            errors.append(f"relationships[{i}] 缺少 to_sheet")
        else:
            validate_sheet_exists(to_sheet, known_sheets, errors, f"relationships[{i}].to_sheet")

        if not keys:
            errors.append(f"relationships[{i}] 至少需要一个 keys")
        else:
            for key_idx, key_item in enumerate(keys, start=1):
                from_field = key_item.get("from_field")
                to_field = key_item.get("to_field")

                if not from_field or not to_field:
                    errors.append(f"relationships[{i}].keys[{key_idx}] 缺少 from_field 或 to_field")
                    continue

                if from_sheet in known_sheets:
                    validate_fields_exist(
                        from_sheet,
                        [from_field],
                        field_index,
                        errors,
                        f"relationships[{i}].keys[{key_idx}].from_field"
                    )
                if to_sheet in known_sheets:
                    validate_fields_exist(
                        to_sheet,
                        [to_field],
                        field_index,
                        errors,
                        f"relationships[{i}].keys[{key_idx}].to_field"
                    )

        if to_sheet in known_sheets and output_fields:
            validate_fields_exist(
                to_sheet,
                output_fields,
                field_index,
                errors,
                f"relationships[{i}].output_fields"
            )

    available_fields = set(field_index.get(main_fact_sheet, set())) if main_fact_sheet else set()

    for rel in relationships:
        for field in rel.get("output_fields", []):
            if field:
                available_fields.add(field)

    for i, rule in enumerate(rules, start=1):
        source_sheet = rule.get("source_sheet")
        rule_type = rule.get("rule_type")
        match_keys = rule.get("match_keys", [])
        output_field = rule.get("output_field")
        effective_start_field = rule.get("effective_start_field")
        effective_end_field = rule.get("effective_end_field")
        threshold_min_field = rule.get("threshold_min_field")
        threshold_max_field = rule.get("threshold_max_field")
        threshold_base_field = rule.get("threshold_base_field")
        ladder_flag_field = rule.get("ladder_flag_field")
        extra_output_fields = rule.get("extra_output_fields", [])

        if not rule.get("rule_name"):
            errors.append(f"rules[{i}] 缺少 rule_name")

        if not source_sheet:
            errors.append(f"rules[{i}] 缺少 source_sheet")
        else:
            validate_sheet_exists(source_sheet, known_sheets, errors, f"rules[{i}].source_sheet")

        if not rule_type:
            errors.append(f"rules[{i}] 缺少 rule_type")

        if not match_keys:
            warnings.append(f"rules[{i}] 未提供 match_keys")

        if source_sheet in known_sheets:
            validate_fields_exist(
                source_sheet,
                [x for x in match_keys if x],
                field_index,
                errors,
                f"rules[{i}].match_keys"
            )

            if output_field:
                validate_fields_exist(
                    source_sheet,
                    [output_field],
                    field_index,
                    errors,
                    f"rules[{i}].output_field"
                )

            extra_fields = [x for x in [effective_start_field, effective_end_field, ladder_flag_field] if x]
            if extra_fields:
                validate_fields_exist(
                    source_sheet,
                    extra_fields,
                    field_index,
                    errors,
                    f"rules[{i}].effective_or_flag_fields"
                )

            ladder_fields = [x for x in [threshold_min_field, threshold_max_field, threshold_base_field] if x]
            if ladder_fields:
                validate_fields_exist(
                    source_sheet,
                    ladder_fields,
                    field_index,
                    errors,
                    f"rules[{i}].threshold_fields"
                )

            if extra_output_fields:
                validate_fields_exist(
                    source_sheet,
                    [x for x in extra_output_fields if x],
                    field_index,
                    errors,
                    f"rules[{i}].extra_output_fields"
                )

        if rule_type == "effective_lookup":
            if not effective_start_field or not effective_end_field:
                errors.append(f"rules[{i}] rule_type=effective_lookup 时必须提供 effective_start_field 和 effective_end_field")

        if rule_type in {"ladder_lookup", "productline_rate_lookup"}:
            if not output_field:
                errors.append(f"rules[{i}] {rule_type} 时必须提供 output_field")

        if output_field:
            available_fields.add(output_field)
        for f in extra_output_fields:
            if f:
                available_fields.add(f)

    for i, formula in enumerate(formulas, start=1):
        output_field = formula.get("output_field")
        expression = formula.get("expression")

        if not output_field:
            errors.append(f"formulas[{i}] 缺少 output_field")
        if not expression:
            errors.append(f"formulas[{i}] 缺少 expression")

        if expression:
            tokens = re.findall(r"[\u4e00-\u9fffA-Za-z_][\u4e00-\u9fffA-Za-z0-9_]*", expression)
            reserved = {"and", "or", "not", "if", "else", "True", "False", "None"}
            referenced_fields = [t for t in tokens if t not in reserved]

            unknown_fields = [
                f for f in referenced_fields
                if f not in available_fields and not re.fullmatch(r"\d+(\.\d+)?", f)
            ]
            if unknown_fields:
                warnings.append(
                    f"formulas[{i}] 表达式引用了当前未确认字段: {', '.join(sorted(set(unknown_fields)))}"
                )

        if output_field and output_field in available_fields:
            warnings.append(f"formulas[{i}] output_field={output_field} 与已有字段重名，执行时可能会覆盖原值")

        if output_field:
            available_fields.add(output_field)

    return {
        "is_valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings
    }


# =========================
# 执行工具
# =========================
def load_workbook_frames(file_bytes: bytes) -> Dict[str, pd.DataFrame]:
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    frames: Dict[str, pd.DataFrame] = {}

    for ws in wb.worksheets:
        header, data_rows = detect_header_and_rows(ws)
        clean_header = [safe_str(h) for h in header if safe_str(h)]

        if not clean_header:
            frames[ws.title] = pd.DataFrame()
            continue

        keep_indices = [i for i, h in enumerate(header) if safe_str(h)]

        normalized_rows = []
        for row in data_rows:
            row = list(row)
            selected = []
            for idx in keep_indices:
                selected.append(row[idx] if idx < len(row) else None)
            normalized_rows.append([normalize_cell_value(v) for v in selected])

        df = pd.DataFrame(normalized_rows, columns=clean_header)
        frames[ws.title] = df

    return frames


def normalize_join_key_series(s: pd.Series) -> pd.Series:
    return s.fillna("").astype(str).str.strip()


def try_parse_period(v: Any) -> Any:
    if v is None or safe_str(v) == "":
        return pd.NaT

    if isinstance(v, (datetime, date)):
        return pd.to_datetime(v)

    s = safe_str(v)
    patterns = ["%Y-%m", "%Y/%m", "%Y-%m-%d", "%Y/%m/%d"]
    for fmt in patterns:
        try:
            return pd.to_datetime(datetime.strptime(s, fmt))
        except Exception:
            pass

    try:
        return pd.to_datetime(s)
    except Exception:
        return pd.NaT


def ensure_datetime_column(df: pd.DataFrame, col: Optional[str]) -> pd.Series:
    if not col or col not in df.columns:
        return pd.Series([pd.NaT] * len(df), index=df.index)
    return df[col].apply(try_parse_period)


def infer_result_semantic_type(col_name: str, data_type: str) -> str:
    name = normalize_name(col_name)

    if any(k in name for k in ["年月", "日期", "时间", "month", "date"]):
        return "time"
    if any(k in name for k in ["销售", "员工", "姓名"]):
        return "person"
    if any(k in name for k in ["客户", "客户名称"]):
        return "customer"
    if any(k in name for k in ["提点率", "费率", "rate"]):
        return "rate"
    if any(k in name for k in ["金额", "收入", "奖金", "基数", "系数"]):
        return "metric"
    if data_type == "number":
        return "metric"
    return "dimension"


def infer_schema_from_df(df: pd.DataFrame) -> List[Dict[str, Any]]:
    schema = []
    for col in df.columns:
        series = df[col]
        non_null = series.dropna()

        if pd.api.types.is_numeric_dtype(series):
            data_type = "number"
        elif len(non_null) > 0 and non_null.astype(str).str.match(r"^\d{4}-\d{1,2}(-\d{1,2})?$").mean() >= 0.6:
            data_type = "date"
        else:
            data_type = "string"

        semantic_type = infer_result_semantic_type(col, data_type)
        schema.append({
            "name": col,
            "type": data_type,
            "semantic_type": semantic_type
        })
    return schema


def build_metrics_and_dimensions(schema: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[str]]:
    metrics = []
    dimensions = []

    for item in schema:
        name = item["name"]
        data_type = item["type"]
        semantic_type = item["semantic_type"]

        if data_type == "number":
            metrics.append({
                "name": name,
                "supported_agg": ["sum", "avg", "max", "min"]
            })
        else:
            if name not in dimensions:
                dimensions.append(name)

        if semantic_type in {"time", "person", "customer", "dimension"} and data_type != "number":
            if name not in dimensions:
                dimensions.append(name)

    return metrics, dimensions


def replace_formula_fields(expression: str, columns: List[str]) -> str:
    sorted_cols = sorted(columns, key=len, reverse=True)
    expr = expression

    for col in sorted_cols:
        pattern = rf"(?<![A-Za-z0-9_\u4e00-\u9fff]){re.escape(col)}(?![A-Za-z0-9_\u4e00-\u9fff])"
        expr = re.sub(pattern, f"df[{col!r}]", expr)

    return expr


def safe_formula_eval(df: pd.DataFrame, expression: str) -> pd.Series:
    expr = replace_formula_fields(expression, list(df.columns))
    allowed_globals = {
        "__builtins__": {},
        "df": df,
        "pd": pd,
        "np": np,
        "math": math,
    }
    return eval(expr, allowed_globals, {})


def filter_by_ladder_bounds(
        df: pd.DataFrame,
        base_value: Any,
        threshold_min_field: Optional[str],
        threshold_max_field: Optional[str],
) -> pd.DataFrame:
    numeric_base = pd.to_numeric(pd.Series([base_value]), errors="coerce").iloc[0]
    if pd.isna(numeric_base):
        return df.iloc[0:0].copy()

    min_series = (
        pd.to_numeric(df[threshold_min_field], errors="coerce")
        if threshold_min_field and threshold_min_field in df.columns
        else pd.Series([np.nan] * len(df), index=df.index)
    )
    max_series = (
        pd.to_numeric(df[threshold_max_field], errors="coerce")
        if threshold_max_field and threshold_max_field in df.columns
        else pd.Series([np.nan] * len(df), index=df.index)
    )

    # 业务规则：
    # 1. 有阶梯上限：base > 上限 才命中
    # 2. 有阶梯下限：base <= 下限 才命中
    cond_upper = max_series.isna() | (numeric_base > max_series)
    cond_lower = min_series.isna() | (numeric_base <= min_series)

    return df[cond_upper & cond_lower]


def execute_lookup(
        fact_df: pd.DataFrame,
        frames: Dict[str, pd.DataFrame],
        step_config: Dict[str, Any],
        lineage: List[Dict[str, Any]],
) -> pd.DataFrame:
    to_sheet = step_config["to_sheet"]
    if to_sheet not in frames:
        raise ValueError(f"lookup 目标 sheet 不存在: {to_sheet}")

    dim_df = frames[to_sheet].copy()
    keys = step_config.get("keys", [])
    output_fields = [x for x in step_config.get("output_fields", []) if safe_str(x)]

    if not keys:
        raise ValueError("lookup 缺少 keys")

    left_on = [k["from_field"] for k in keys]
    right_on = [k["to_field"] for k in keys]

    left = fact_df.copy()
    right = dim_df.copy()

    for col in left_on:
        if col not in left.columns:
            raise ValueError(f"事实表缺少关联字段: {col}")
        left[col] = normalize_join_key_series(left[col])

    for col in right_on:
        if col not in right.columns:
            raise ValueError(f"维表缺少关联字段: {col}")
        right[col] = normalize_join_key_series(right[col])

    use_cols = [c for c in list(dict.fromkeys(right_on + output_fields)) if c in right.columns]

    merged = left.merge(
        right[use_cols],
        how="left",
        left_on=left_on,
        right_on=right_on,
        suffixes=("", "_dim")
    )

    for rcol in right_on:
        if rcol in merged.columns and rcol not in fact_df.columns:
            merged.drop(columns=[rcol], inplace=True, errors="ignore")

    logic = " + ".join([f"{k['from_field']}={k['to_field']}" for k in keys])
    for field in output_fields:
        lineage.append({
            "field": field,
            "source_sheet": to_sheet,
            "logic": logic
        })

    return merged


def execute_effective_lookup(
        fact_df: pd.DataFrame,
        frames: Dict[str, pd.DataFrame],
        rule: Dict[str, Any],
        lineage: List[Dict[str, Any]],
) -> pd.DataFrame:
    source_sheet = rule["source_sheet"]
    if source_sheet not in frames:
        raise ValueError(f"effective_lookup 源 sheet 不存在: {source_sheet}")

    rule_df = frames[source_sheet].copy()

    match_keys = rule.get("match_keys", [])
    date_field = rule.get("date_field")
    start_field = rule.get("effective_start_field")
    end_field = rule.get("effective_end_field")
    output_field = rule.get("output_field")

    if not output_field:
        raise ValueError("effective_lookup 缺少 output_field")

    ensure_object_column(fact_df, output_field)

    fact_dates = ensure_datetime_column(fact_df, date_field)
    rule_start = ensure_datetime_column(rule_df, start_field)
    rule_end = ensure_datetime_column(rule_df, end_field)

    for idx in fact_df.index:
        matched = rule_df.copy()

        for key in match_keys:
            if key not in fact_df.columns or key not in matched.columns:
                continue
            matched = matched[
                normalize_join_key_series(matched[key]) ==
                safe_str(fact_df.at[idx, key])
                ]

        current_date = fact_dates.loc[idx]
        if pd.notna(current_date) and start_field and end_field:
            matched = matched[(rule_start <= current_date) & (rule_end >= current_date)]

        if len(matched) > 0 and output_field in matched.columns:
            fact_df.at[idx, output_field] = matched.iloc[0][output_field]

    logic_parts = []
    if match_keys:
        logic_parts.append(" + ".join(match_keys))
    if start_field and end_field:
        logic_parts.append("生效失效区间")

    lineage.append({
        "field": output_field,
        "source_sheet": source_sheet,
        "logic": " + ".join(logic_parts) if logic_parts else "规则匹配"
    })

    return fact_df


def execute_ladder_lookup(
        fact_df: pd.DataFrame,
        frames: Dict[str, pd.DataFrame],
        rule: Dict[str, Any],
        lineage: List[Dict[str, Any]],
) -> pd.DataFrame:
    source_sheet = rule["source_sheet"]
    if source_sheet not in frames:
        raise ValueError(f"ladder_lookup 源 sheet 不存在: {source_sheet}")

    rule_df = frames[source_sheet].copy()

    match_keys = rule.get("match_keys", [])
    date_field = rule.get("date_field")
    start_field = rule.get("effective_start_field")
    end_field = rule.get("effective_end_field")
    threshold_min_field = rule.get("threshold_min_field")
    threshold_max_field = rule.get("threshold_max_field")
    threshold_base_field = rule.get("threshold_base_field")
    output_field = rule.get("output_field")

    if not output_field:
        raise ValueError("ladder_lookup 缺少 output_field")

    if not threshold_min_field and not threshold_max_field:
        raise ValueError("ladder_lookup 至少需要 threshold_min_field 或 threshold_max_field")

    ensure_object_column(fact_df, output_field)

    fact_dates = ensure_datetime_column(fact_df, date_field)
    rule_start = ensure_datetime_column(rule_df, start_field)
    rule_end = ensure_datetime_column(rule_df, end_field)

    if not threshold_base_field or threshold_base_field not in fact_df.columns:
        candidates = ["奖金计算基数", "业绩", "金额", "收入", "base"]
        for c in candidates:
            if c in fact_df.columns:
                threshold_base_field = c
                break

    if not threshold_base_field or threshold_base_field not in fact_df.columns:
        raise ValueError("ladder_lookup 无法确定 threshold_base_field，请在 rule config 中显式提供")

    for idx in fact_df.index:
        matched = rule_df.copy()

        for key in match_keys:
            if key not in fact_df.columns or key not in matched.columns:
                continue
            matched = matched[
                normalize_join_key_series(matched[key]) ==
                safe_str(fact_df.at[idx, key])
                ]

        current_date = fact_dates.loc[idx]
        if pd.notna(current_date) and start_field and end_field:
            matched = matched[(rule_start <= current_date) & (rule_end >= current_date)]

        matched = filter_by_ladder_bounds(
            matched,
            fact_df.at[idx, threshold_base_field],
            threshold_min_field,
            threshold_max_field,
        )

        if len(matched) > 0 and output_field in matched.columns:
            fact_df.at[idx, output_field] = matched.iloc[0][output_field]

    logic_parts = []
    if match_keys:
        logic_parts.append(" + ".join(match_keys))
    if start_field and end_field:
        logic_parts.append("生效失效区间")
    if threshold_max_field:
        logic_parts.append(f"{threshold_base_field} > {threshold_max_field}")
    if threshold_min_field:
        logic_parts.append(f"{threshold_base_field} <= {threshold_min_field}")

    lineage.append({
        "field": output_field,
        "source_sheet": source_sheet,
        "logic": " + ".join(logic_parts)
    })

    return fact_df


def execute_productline_rate_lookup(
        fact_df: pd.DataFrame,
        frames: Dict[str, pd.DataFrame],
        rule: Dict[str, Any],
        lineage: List[Dict[str, Any]],
) -> pd.DataFrame:
    source_sheet = rule["source_sheet"]
    if source_sheet not in frames:
        raise ValueError(f"productline_rate_lookup 源 sheet 不存在: {source_sheet}")

    rule_df = frames[source_sheet].copy()

    match_keys = rule.get("match_keys", [])
    date_field = rule.get("date_field")
    start_field = rule.get("effective_start_field")
    end_field = rule.get("effective_end_field")
    ladder_flag_field = rule.get("ladder_flag_field")
    threshold_min_field = rule.get("threshold_min_field")
    threshold_max_field = rule.get("threshold_max_field")
    threshold_base_field = rule.get("threshold_base_field")
    output_field = rule.get("output_field")
    extra_output_fields = rule.get("extra_output_fields", []) or []

    if not output_field:
        raise ValueError("productline_rate_lookup 缺少 output_field")

    for col in [output_field] + extra_output_fields:
        if col:
            ensure_object_column(fact_df, col)

    fact_dates = ensure_datetime_column(fact_df, date_field)
    rule_start = ensure_datetime_column(rule_df, start_field) if start_field else pd.Series([pd.NaT] * len(rule_df), index=rule_df.index)
    rule_end = ensure_datetime_column(rule_df, end_field) if end_field else pd.Series([pd.NaT] * len(rule_df), index=rule_df.index)

    if not threshold_base_field or threshold_base_field not in fact_df.columns:
        for c in ["奖金计算基数", "业绩", "金额", "收入", "base"]:
            if c in fact_df.columns:
                threshold_base_field = c
                break

    for idx in fact_df.index:
        matched = rule_df.copy()

        # 1. 先按产品线名称等维度匹配
        for key in match_keys:
            if key not in fact_df.columns or key not in matched.columns:
                continue
            matched = matched[
                normalize_join_key_series(matched[key]) ==
                safe_str(fact_df.at[idx, key])
                ]

        # 2. 再按生效失效区间过滤
        current_date = fact_dates.loc[idx]
        if pd.notna(current_date) and start_field and end_field:
            matched = matched[(rule_start <= current_date) & (rule_end >= current_date)]

        if matched.empty:
            continue

        selected = None

        # 3. 判断是否阶梯提点
        if ladder_flag_field and ladder_flag_field in matched.columns:
            ladder_rows = matched[matched[ladder_flag_field].apply(truthy_flag)]
            non_ladder_rows = matched[~matched[ladder_flag_field].apply(truthy_flag)]
        else:
            ladder_rows = matched
            non_ladder_rows = pd.DataFrame(columns=matched.columns)

        # 4. 若为阶梯提点，则按新的上限/下限规则命中
        if not ladder_rows.empty and threshold_base_field and threshold_base_field in fact_df.columns:
            ladder_hit = filter_by_ladder_bounds(
                ladder_rows,
                fact_df.at[idx, threshold_base_field],
                threshold_min_field,
                threshold_max_field,
            )
            if not ladder_hit.empty:
                selected = ladder_hit.iloc[0]

        # 5. 若不是阶梯或阶梯没命中，则取非阶梯记录
        if selected is None and not non_ladder_rows.empty:
            selected = non_ladder_rows.iloc[0]

        # 6. 最后兜底
        if selected is None and not matched.empty:
            selected = matched.iloc[0]

        if selected is None:
            continue

        if output_field in selected.index:
            fact_df.at[idx, output_field] = selected[output_field]

        for col in extra_output_fields:
            if col in selected.index:
                fact_df.at[idx, col] = selected[col]

    logic_parts = []
    if match_keys:
        logic_parts.append(f"按{'+'.join(match_keys)}匹配产品线规则")
    if start_field and end_field:
        logic_parts.append("按生效失效区间过滤")
    if ladder_flag_field:
        logic_parts.append(f"依据{ladder_flag_field}判断是否阶梯提点")
    if threshold_max_field and threshold_base_field:
        logic_parts.append(f"若有{threshold_max_field}则要求{threshold_base_field}>{threshold_max_field}")
    if threshold_min_field and threshold_base_field:
        logic_parts.append(f"若有{threshold_min_field}则要求{threshold_base_field}<={threshold_min_field}")
    logic_parts.append(f"输出{output_field}")

    lineage.append({
        "field": output_field,
        "source_sheet": source_sheet,
        "logic": "；".join(logic_parts)
    })

    for col in extra_output_fields:
        lineage.append({
            "field": col,
            "source_sheet": source_sheet,
            "logic": f"按产品线规则回填{col}"
        })

    return fact_df


def execute_formula(
        fact_df: pd.DataFrame,
        formula: Dict[str, Any],
        lineage: List[Dict[str, Any]],
) -> pd.DataFrame:
    output_field = formula["output_field"]
    expression = formula["expression"]

    result = safe_formula_eval(fact_df, expression)
    fact_df[output_field] = result

    source_fields = [c for c in fact_df.columns if c in expression]

    lineage.append({
        "field": output_field,
        "source_fields": source_fields,
        "logic": expression
    })

    return fact_df


def dataframe_preview(df: pd.DataFrame, limit: int = 20) -> List[Dict[str, Any]]:
    preview_df = df.head(limit).copy()
    preview_df = preview_df.where(pd.notnull(preview_df), None)
    records = preview_df.to_dict(orient="records")
    return to_jsonable(records)


def compute_stats(df: pd.DataFrame) -> Dict[str, Any]:
    stats: Dict[str, Any] = {}

    if "应发奖金" in df.columns:
        bonus = pd.to_numeric(df["应发奖金"], errors="coerce")
        if bonus.notna().any():
            stats["total_bonus"] = float(round(bonus.sum(), 2))
            stats["avg_bonus"] = float(round(bonus.mean(), 2))

    return stats


# =========================
# 查询工具
# =========================
def apply_filters(df: pd.DataFrame, filters: Dict[str, Any]) -> pd.DataFrame:
    if not filters:
        return df

    filtered = df.copy()

    for field, expected in filters.items():
        if field not in filtered.columns:
            continue

        if isinstance(expected, list):
            filtered = filtered[filtered[field].isin(expected)]
        else:
            filtered = filtered[
                filtered[field].fillna("").astype(str).str.strip() ==
                safe_str(expected)
                ]

    return filtered


def parse_sort_by(sort_by: List[str]) -> Tuple[List[str], List[bool]]:
    cols: List[str] = []
    ascending: List[bool] = []

    for item in sort_by or []:
        s = safe_str(item)
        if not s:
            continue

        parts = s.split()
        col = parts[0]
        desc = len(parts) > 1 and parts[1].lower() == "desc"
        cols.append(col)
        ascending.append(not desc)

    return cols, ascending


def select_existing_columns(df: pd.DataFrame, cols: List[str]) -> List[str]:
    return [c for c in cols if c in df.columns]


def aggregate_df(df: pd.DataFrame, group_by: List[str], metrics: List[str]) -> pd.DataFrame:
    valid_group_by = select_existing_columns(df, group_by or [])
    valid_metrics = select_existing_columns(df, metrics or [])

    if not valid_metrics:
        return df.copy()

    agg_map = {m: "sum" for m in valid_metrics}

    temp_df = df.copy()
    for m in valid_metrics:
        temp_df[m] = pd.to_numeric(temp_df[m], errors="coerce")

    if valid_group_by:
        result = temp_df.groupby(valid_group_by, dropna=False, as_index=False).agg(agg_map)
    else:
        agg_row = {m: temp_df[m].sum() for m in valid_metrics}
        result = pd.DataFrame([agg_row])

    return result


def build_breakdown_from_row(row: Dict[str, Any], preferred_fields: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    if preferred_fields is None:
        preferred_fields = ["奖金计算基数", "调整系数", "提点率a", "提点率b", "应发奖金"]

    breakdown = []
    for field in preferred_fields:
        if field in row:
            breakdown.append({
                "field": field,
                "value": row.get(field)
            })
    return breakdown


def get_dataset_record(dataset_id: str) -> Dict[str, Any]:
    record = DATASET_STORE.get(dataset_id)
    if not record:
        raise HTTPException(status_code=404, detail=f"未找到 dataset_id: {dataset_id}")
    return record


def get_lineage_map(dataset_record: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    lineage_list = dataset_record.get("result", {}).get("lineage", []) or []
    lineage_map: Dict[str, Dict[str, Any]] = {}

    for item in lineage_list:
        field = item.get("field")
        if field:
            lineage_map[field] = item

    return lineage_map


def normalize_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return to_jsonable(records)


def query_metric(
        df: pd.DataFrame,
        payload: Dict[str, Any],
        dataset_record: Dict[str, Any],
) -> Dict[str, Any]:
    metrics = payload.get("metrics", []) or []
    need_rows = bool(payload.get("need_rows", True))
    need_breakdown = bool(payload.get("need_breakdown", True))
    need_lineage = bool(payload.get("need_lineage", False))
    filters = payload.get("filters", {}) or {}

    filtered = apply_filters(df, filters)

    summary: Dict[str, Any] = {}
    for m in metrics:
        if m in filtered.columns:
            summary[m] = float(pd.to_numeric(filtered[m], errors="coerce").sum())

    rows: List[Dict[str, Any]] = []
    if need_rows:
        keep_cols = []
        if not filtered.empty:
            base_cols = ["统计年月", "销售", "客户名称"] + metrics
            keep_cols = [c for c in base_cols if c in filtered.columns]
            if not keep_cols:
                keep_cols = list(filtered.columns[:10])
        rows = normalize_records(filtered[keep_cols].head(100).to_dict(orient="records")) if keep_cols else []

    breakdown: List[Dict[str, Any]] = []
    if need_breakdown and not filtered.empty:
        first_row = to_jsonable(filtered.iloc[0].to_dict())
        breakdown = build_breakdown_from_row(first_row)

    response = {
        "dataset_id": payload["dataset_id"],
        "intent": "metric_query",
        "summary": to_jsonable(summary),
        "rows": rows,
        "breakdown": breakdown,
        "warnings": []
    }

    if need_lineage:
        lineage_map = get_lineage_map(dataset_record)
        response["lineage"] = [lineage_map[m] for m in metrics if m in lineage_map]

    return response


def query_ranking(df: pd.DataFrame, payload: Dict[str, Any]) -> Dict[str, Any]:
    metrics = payload.get("metrics", []) or []
    group_by = payload.get("group_by", []) or []
    filters = payload.get("filters", {}) or {}
    sort_by = payload.get("sort_by", []) or []
    top_n = payload.get("top_n")
    need_rows = bool(payload.get("need_rows", True))

    filtered = apply_filters(df, filters)
    result_df = aggregate_df(filtered, group_by, metrics)

    sort_cols, ascending = parse_sort_by(sort_by)
    valid_sort_cols = [c for c in sort_cols if c in result_df.columns]
    if valid_sort_cols:
        valid_ascending = []
        for c in valid_sort_cols:
            i = sort_cols.index(c)
            valid_ascending.append(ascending[i])
        result_df = result_df.sort_values(valid_sort_cols, ascending=valid_ascending, na_position="last")
    elif metrics:
        first_metric = metrics[0]
        if first_metric in result_df.columns:
            result_df = result_df.sort_values(first_metric, ascending=False, na_position="last")

    if top_n:
        result_df = result_df.head(int(top_n))

    rows = []
    if need_rows:
        keep_cols = select_existing_columns(result_df, group_by + metrics)
        rows = normalize_records(result_df[keep_cols].to_dict(orient="records")) if keep_cols else []

    return {
        "dataset_id": payload["dataset_id"],
        "intent": "ranking",
        "summary": {
            "top_n": top_n,
            "metric": metrics[0] if metrics else None
        },
        "rows": rows,
        "warnings": []
    }


def query_trend(df: pd.DataFrame, payload: Dict[str, Any]) -> Dict[str, Any]:
    metrics = payload.get("metrics", []) or []
    filters = payload.get("filters", {}) or {}
    group_by = payload.get("group_by", []) or []

    filtered = apply_filters(df, filters)

    time_field = None
    for c in ["统计年月", "年月", "日期"]:
        if c in filtered.columns:
            time_field = c
            break

    if not time_field:
        return {
            "dataset_id": payload["dataset_id"],
            "intent": "trend",
            "summary": {},
            "rows": [],
            "warnings": ["未找到时间字段，无法进行趋势分析"]
        }

    trend_group_by = [time_field] + [c for c in group_by if c != time_field]
    result_df = aggregate_df(filtered, trend_group_by, metrics)
    result_df = result_df.sort_values(time_field)

    keep_cols = select_existing_columns(result_df, trend_group_by + metrics)
    return {
        "dataset_id": payload["dataset_id"],
        "intent": "trend",
        "summary": {
            "time_field": time_field,
            "metric": metrics[0] if metrics else None
        },
        "rows": normalize_records(result_df[keep_cols].to_dict(orient="records")) if keep_cols else [],
        "warnings": []
    }


def query_compare(df: pd.DataFrame, payload: Dict[str, Any]) -> Dict[str, Any]:
    metrics = payload.get("metrics", []) or []
    filters = payload.get("filters", {}) or {}
    group_by = payload.get("group_by", []) or []

    filtered = apply_filters(df, filters)
    result_df = aggregate_df(filtered, group_by, metrics)

    keep_cols = select_existing_columns(result_df, group_by + metrics)
    return {
        "dataset_id": payload["dataset_id"],
        "intent": "compare",
        "summary": {
            "metrics": metrics,
            "group_by": group_by
        },
        "rows": normalize_records(result_df[keep_cols].to_dict(orient="records")) if keep_cols else [],
        "warnings": []
    }


def query_breakdown(df: pd.DataFrame, payload: Dict[str, Any]) -> Dict[str, Any]:
    metrics = payload.get("metrics", []) or []
    filters = payload.get("filters", {}) or {}
    group_by = payload.get("group_by", []) or []

    filtered = apply_filters(df, filters)
    result_df = aggregate_df(filtered, group_by, metrics)

    keep_cols = select_existing_columns(result_df, group_by + metrics)
    return {
        "dataset_id": payload["dataset_id"],
        "intent": "breakdown",
        "summary": {
            "metrics": metrics,
            "group_by": group_by
        },
        "rows": normalize_records(result_df[keep_cols].to_dict(orient="records")) if keep_cols else [],
        "warnings": []
    }


def query_lineage_explain(dataset_record: Dict[str, Any], payload: Dict[str, Any]) -> Dict[str, Any]:
    fields = payload.get("fields", []) or []
    lineage_map = get_lineage_map(dataset_record)

    lineage = []
    for field in fields:
        if field in lineage_map:
            lineage.append(lineage_map[field])
        else:
            lineage.append({
                "field": field,
                "logic": "未找到字段血缘信息"
            })

    return {
        "dataset_id": payload["dataset_id"],
        "intent": "lineage_explain",
        "lineage": to_jsonable(lineage),
        "warnings": []
    }


def query_rule_explain(dataset_record: Dict[str, Any], payload: Dict[str, Any]) -> Dict[str, Any]:
    fields = payload.get("fields", []) or []
    lineage_map = get_lineage_map(dataset_record)

    explanations = []
    for field in fields:
        item = lineage_map.get(field)
        if item:
            explanations.append({
                "field": field,
                "logic": item.get("logic"),
                "source_sheet": item.get("source_sheet")
            })
        else:
            explanations.append({
                "field": field,
                "logic": "未找到规则说明"
            })

    return {
        "dataset_id": payload["dataset_id"],
        "intent": "rule_explain",
        "rules": to_jsonable(explanations),
        "warnings": []
    }


# =========================
# 接口 1：解析工作簿
# =========================
@app.post("/v1/workbooks/parse")
async def parse_workbook(
        file: UploadFile = File(...),
        instruction: Optional[str] = Form(None),
        detect_relationships: bool = Form(True),
        infer_rules_flag: bool = Form(True),
        preview_rows: int = Form(20),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="缺少上传文件")

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="仅支持 xlsx/xlsm/xltx/xltm 文件")

    try:
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败: {str(e)}")

    request_id = gen_request_id()
    file_id = gen_file_id()
    FILE_STORE[file_id] = content

    warnings: List[str] = []

    try:
        sheet_profiles: List[Dict[str, Any]] = []
        instruction_extract: Dict[str, Any] = {
            "has_instruction_sheet": False,
            "sheet_name": None,
            "text_summary": None
        }

        for ws in wb.worksheets:
            profile = build_sheet_profile(ws, preview_rows=preview_rows)
            sheet_profiles.append(profile)

            if profile["candidate_role"] == "instruction":
                instruction_extract["has_instruction_sheet"] = True
                instruction_extract["sheet_name"] = ws.title
                instruction_extract["text_summary"] = summarize_instruction_sheet(profile["_data_rows"])

        candidate_relationships: List[Dict[str, Any]] = []
        if detect_relationships:
            candidate_relationships = infer_relationships(sheet_profiles)

        candidate_rules: List[Dict[str, Any]] = []
        if infer_rules_flag:
            candidate_rules = infer_candidate_rules(sheet_profiles)

        if instruction:
            warnings.append(f"已接收业务说明 instruction: {instruction[:80]}")

        response = {
            "request_id": request_id,
            "workbook": {
                "file_id": file_id,
                "file_name": file.filename,
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames
            },
            "sheets": cleanup_internal_fields(sheet_profiles),
            "candidate_relationships": candidate_relationships,
            "candidate_rules": candidate_rules,
            "instruction_extract": instruction_extract,
            "warnings": warnings
        }

        response = to_jsonable(response)
        PARSED_WORKBOOK_STORE[file_id] = response
        return response

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"parse_workbook 处理失败: {str(e)}")


# =========================
# 接口 2：构建数据模型
# =========================
@app.post("/v1/models/build")
async def build_model(payload: Dict[str, Any]):
    file_id = payload.get("file_id")
    if not file_id:
        raise HTTPException(status_code=400, detail="缺少 file_id")

    parsed_workbook = PARSED_WORKBOOK_STORE.get(file_id)
    if not parsed_workbook:
        raise HTTPException(status_code=404, detail=f"未找到 file_id 对应的解析结果: {file_id}")

    validation = validate_model_request(payload, parsed_workbook)

    main_fact_sheet = payload.get("main_fact_sheet")
    relationships = payload.get("relationships", [])
    rules = payload.get("rules", [])
    formulas = payload.get("formulas", [])

    plan_id = gen_plan_id()
    model_id = gen_model_id()

    steps = build_plan_steps(
        relationships=relationships,
        rules=rules,
        formulas=formulas,
    )

    plan = {
        "plan_id": plan_id,
        "main_fact_sheet": main_fact_sheet,
        "steps": steps
    }

    response = {
        "model_id": model_id,
        "plan": plan,
        "validation": validation
    }

    response = to_jsonable(response)

    MODEL_STORE[model_id] = {
        "model_id": model_id,
        "file_id": file_id,
        "request": deepcopy(payload),
        "response": deepcopy(response)
    }

    return response


# =========================
# 接口 3：执行计算
# =========================
@app.post("/v1/calculations/execute")
async def execute_calculation(payload: Dict[str, Any]):
    file_id = payload.get("file_id")
    model_id = payload.get("model_id")
    plan = payload.get("plan")
    execution_options = payload.get("execution_options", {})

    if not file_id:
        raise HTTPException(status_code=400, detail="缺少 file_id")

    if file_id not in FILE_STORE:
        raise HTTPException(status_code=404, detail=f"未找到 file_id 对应的原始文件: {file_id}")

    if not model_id and not plan:
        raise HTTPException(status_code=400, detail="model_id 和 plan 至少传一个")

    if model_id:
        model_record = MODEL_STORE.get(model_id)
        if not model_record:
            raise HTTPException(status_code=404, detail=f"未找到 model_id: {model_id}")

        model_response = model_record.get("response", {})
        plan = model_response.get("plan")

    if not plan:
        raise HTTPException(status_code=400, detail="无法解析执行 plan")

    main_fact_sheet = plan.get("main_fact_sheet")
    steps = plan.get("steps", [])

    file_bytes = FILE_STORE[file_id]
    frames = load_workbook_frames(file_bytes)

    if main_fact_sheet not in frames:
        raise HTTPException(status_code=400, detail=f"主表不存在: {main_fact_sheet}")

    fact_df = frames[main_fact_sheet].copy()
    lineage: List[Dict[str, Any]] = []
    warnings: List[str] = []

    for step in steps:
        action = step.get("action")
        config = step.get("config", {})

        try:
            if action == "lookup":
                fact_df = execute_lookup(fact_df, frames, config, lineage)
            elif action == "effective_lookup":
                fact_df = execute_effective_lookup(fact_df, frames, config, lineage)
            elif action == "ladder_lookup":
                fact_df = execute_ladder_lookup(fact_df, frames, config, lineage)
            elif action == "productline_rate_lookup":
                fact_df = execute_productline_rate_lookup(fact_df, frames, config, lineage)
            elif action == "formula":
                fact_df = execute_formula(fact_df, config, lineage)
            else:
                warnings.append(f"未识别的 step action，已跳过: {action}")
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"执行步骤失败 step={step.get('name')} action={action}: {str(e)}"
            )

    schema = infer_schema_from_df(fact_df)
    metrics, dimensions = build_metrics_and_dimensions(schema)
    preview_rows = int(execution_options.get("return_preview_rows", 20))

    dataset_id = gen_dataset_id()
    job_id = gen_job_id()

    result = {
        "job_id": job_id,
        "status": "succeeded",
        "dataset_id": dataset_id,
        "result_table_name": "fact_result",
        "row_count": int(len(fact_df)),
        "schema": schema,
        "metrics": metrics,
        "dimensions": dimensions,
        "lineage": lineage if execution_options.get("materialize_lineage", True) else [],
        "preview": dataframe_preview(fact_df, preview_rows),
        "stats": compute_stats(fact_df),
        "warnings": warnings
    }

    result = to_jsonable(result)

    if execution_options.get("persist_result", True):
        record = {
            "job_id": job_id,
            "dataset_id": dataset_id,
            "file_id": file_id,
            "model_id": model_id,
            "plan": deepcopy(plan),
            "result_df": fact_df.copy(),
            "result": deepcopy(result)
        }
        EXECUTION_STORE[job_id] = record
        DATASET_STORE[dataset_id] = record

    return result


# =========================
# 接口 4：查询结果表
# =========================
@app.post("/v1/datasets/query")
async def query_dataset(payload: Dict[str, Any]):
    dataset_id = payload.get("dataset_id")
    intent = payload.get("intent")

    if not dataset_id:
        raise HTTPException(status_code=400, detail="缺少 dataset_id")
    if not intent:
        raise HTTPException(status_code=400, detail="缺少 intent")

    dataset_record = get_dataset_record(dataset_id)
    df = dataset_record.get("result_df")
    if df is None or not isinstance(df, pd.DataFrame):
        raise HTTPException(status_code=500, detail=f"dataset_id={dataset_id} 对应结果表不存在")

    supported_intents = {
        "metric_query",
        "ranking",
        "trend",
        "compare",
        "breakdown",
        "rule_explain",
        "lineage_explain",
    }

    if intent not in supported_intents:
        raise HTTPException(status_code=400, detail=f"不支持的 intent: {intent}")

    if intent == "metric_query":
        response = query_metric(df, payload, dataset_record)
    elif intent == "ranking":
        response = query_ranking(df, payload)
    elif intent == "trend":
        response = query_trend(df, payload)
    elif intent == "compare":
        response = query_compare(df, payload)
    elif intent == "breakdown":
        response = query_breakdown(df, payload)
    elif intent == "rule_explain":
        response = query_rule_explain(dataset_record, payload)
    elif intent == "lineage_explain":
        response = query_lineage_explain(dataset_record, payload)
    else:
        raise HTTPException(status_code=400, detail=f"不支持的 intent: {intent}")

    return to_jsonable(response)


# =========================
# 健康检查
# =========================
@app.get("/health")
async def health():
    return {"status": "ok"}


# =========================
# 本地启动:
# uvicorn parse_workbook:app --reload --host 0.0.0.0 --port 8000
# =========================
